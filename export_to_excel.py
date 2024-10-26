import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import re

input_file_path = "/Users/jiyuchen/Documents/MDP/deep-text-recognition-benchmark/infer_result.txt"
output_file_path = "/Users/jiyuchen/Documents/MDP/deep-text-recognition-benchmark/output.xlsx"

threshold = 0.95

def get_label(input_string):
    pattern = r'^(1[0-1]|[1-9])([A-E])([NSEW])(\d+)$'
    match = re.match(pattern, input_string)
    if match:
        return match.groups()
    return None

def get_count(input_string):
    pattern = r'^(\d+)(pc|PC)$'
    match = re.match(pattern, input_string)
    if match:
        return match.group(1)
    return None

def is_date(input_string):
    pattern = r'^(1[0-2]|0?[1-9])-(3[01]|[12][0-9]|0?[1-9])$'
    return bool(re.match(pattern, input_string))

def get_type(input_string):
    pattern = r'^(BNSF|BN|KI)(3|4|5|SG|IG)(OAK|M|MIX|MIXED)?$'
    match = re.fullmatch(pattern, input_string)
    if match:
        return match.groups()
    return None

def is_company(input_string):
    pattern = r'^(BNSF|BN|KI)$'
    return bool(re.fullmatch(pattern, input_string))

def get_species(input_string):
    pattern = r'^(3|4|5|SG|IG)(OAK|M|MIX|MIXED)?$'
    match = re.fullmatch(pattern, input_string)
    if match:
        return match.groups()
    return None

if __name__ == "__main__":
    data = {}

    with open(input_file_path, "r") as file:
        for line in file:
            line = line.strip()
            parts = line.split()
            match = re.match(r'.*\/(stack_\d+)_.*', parts[0])
            if match:
                stack_prefix = match.group(1)
                if stack_prefix not in data:
                    data[stack_prefix] = []
                try:
                    value = float(parts[2])
                    data[stack_prefix].append((parts[1], value))
                except ValueError:
                    print(f"Could not convert {parts[2]} to float.")

    data_dict = {}
    label_dict = {}

    for stack, value in data.items():
        has_label = False
        label = ""
        result = ["" for _ in range(11)]
        
        result[10] = stack

        for pair in value:
            entry, confidence = pair
            if get_label(entry):
                has_label = True
                label = entry
                parts = get_label(entry)
                result[0], result[1], result[2] = (parts[0], confidence), (parts[1] + parts[2], confidence), (parts[3], confidence)
            elif is_date(entry):
                result[3] = (entry, confidence)
            elif get_count(entry):
                result[7] = (get_count(entry), confidence)
            elif is_company(entry):
                result[9] = (entry, confidence)
            elif get_species(entry):
                parts = get_species(entry)
                result[5], result[6] = (parts[0], confidence), (parts[1], confidence)
            elif get_type(entry):
                parts = get_type(entry)
                result[5], result[6], result[9] = (parts[1], confidence), (parts[2], confidence), (parts[0], confidence)

        length = sum(1 for val in result if val != "")

        if has_label:
            if label in label_dict:
                if length > label_dict[label]:
                    label_dict[label] = length 
                    data_dict[label] = result
            else:
                label_dict[label] = length
                data_dict[label] = result
     
    formatted_data = []

    for label in data_dict:
        formatted_data.append(data_dict[label])
    
    df_data = [[val[0] if isinstance(val, tuple) else val for val in values] for values in formatted_data]
    
    df = pd.DataFrame(df_data, columns=['TRK', 'SECT', 'ROW', 'DATE', 'XT/SWT', 'Grd/Length', 'Species', 'PIECES', 'NOTE', 'COMPANY', "IMAGE"])

    df.to_excel(output_file_path, index=False)

    wb = openpyxl.load_workbook(output_file_path)
    ws = wb.active

    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for i, values in enumerate(formatted_data, start=2):
        for j, val in enumerate(values, start=1):
            if isinstance(val, tuple) and val[1] < threshold:
                ws.cell(row=i, column=j).fill = highlight_fill

    wb.save(output_file_path)

    print(f"Data has been successfully written to {output_file_path}")