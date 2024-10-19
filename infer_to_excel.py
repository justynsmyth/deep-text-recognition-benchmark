import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import re

# Initialize a dictionary to store the rows of data
data = {}
highlighted_cells = {}

# Define the path to the input file
input_file_path = "/Users/jiyuchen/Documents/MDP/deep-text-recognition-benchmark/infer_result.txt"

# Read the file and process each line
with open(input_file_path, "r") as file:
    for line in file:
        # Strip any leading/trailing whitespace
        line = line.strip()
        
        # Split the line by any whitespace
        parts = line.split()
        
        # Extract the stack prefix (e.g., 'stack_0051')
        match = re.match(r'test_data/(stack_\d+)_.*', parts[0])
        if match:
            stack_prefix = match.group(1)
            
            # Initialize the key in the dictionary if not already present
            if stack_prefix not in data:
                data[stack_prefix] = []
                highlighted_cells[stack_prefix] = []
            
            # Append the values (excluding file path) to the list for this stack
            try:
                value = float(parts[2])
                if value > 0.95:
                    data[stack_prefix].append(parts[1])
                else:
                    data[stack_prefix].append(parts[1])
                    highlighted_cells[stack_prefix].append(len(data[stack_prefix]) - 1)
            except ValueError:
                print(f"Could not convert {parts[2]} to float.")

# Determine the maximum number of columns we'll need
max_columns = max(len(values) for values in data.values())

# Ensure all rows have the same number of columns by padding with empty strings
for key in data:
    data[key].extend([''] * (max_columns - len(data[key])))

# Convert the dictionary to a DataFrame
df = pd.DataFrame.from_dict(data, orient='index')

# Create column headers
columns = []
for i in range(max_columns):
    columns.append(f"Label_{i+1}")

# Assign the generated column names to the DataFrame
df.columns = columns

# Define the path to the output Excel file
output_file_path = "/Users/jiyuchen/Documents/MDP/deep-text-recognition-benchmark/output.xlsx"

# Save the DataFrame to an Excel file
df.to_excel(output_file_path, index_label="StackPrefix")

# Highlight cells in the Excel file that have values <= 0.95
wb = openpyxl.load_workbook(output_file_path)
ws = wb.active

# Define the highlighting style
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Iterate through the dictionary to apply styles
for stack_prefix, columns_to_highlight in highlighted_cells.items():
    row_idx = df.index.get_loc(stack_prefix) + 2  # DataFrame index starts at 2 in excel
    for col_idx in columns_to_highlight:
        excel_column = col_idx + 2  # Adjusting column index for correct excel col (Excel is 1-indexed + StackPrefix column)
        ws.cell(row=row_idx, column=excel_column).fill = highlight_fill

# Save the modified Excel file with styles
wb.save(output_file_path)

print(f"Data has been successfully written to {output_file_path}")